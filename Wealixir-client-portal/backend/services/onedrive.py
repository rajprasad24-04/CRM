import io
import requests
import openpyxl
from config.authentication import (
    ONEDRIVE_TENANT_ID as TENANT_ID, ONEDRIVE_CLIENT_ID as CLIENT_ID,
    ONEDRIVE_CLIENT_SECRET as CLIENT_SECRET, ONEDRIVE_DRIVE_ID as DRIVE_ID,
    ONEDRIVE_FOLDER_NAME as FOLDER_NAME,
)

EXCEL_FILENAME = 'registrations.xlsx'
HEADERS_ROW    = ['ID', 'Name', 'DOB', 'PAN Number', 'Email', 'Mobile', 'Address', 'City', 'Notes', 'Submitted At']
GRAPH = 'https://graph.microsoft.com/v1.0'

def _get_token():
    url = f'https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token'
    resp = requests.post(url, data={'grant_type': 'client_credentials', 'client_id': CLIENT_ID, 'client_secret': CLIENT_SECRET, 'scope': 'https://graph.microsoft.com/.default'})
    resp.raise_for_status()
    return resp.json()['access_token']

def _headers(token):
    return {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}

def _get_or_create_folder(token):
    url = f'{GRAPH}/drives/{DRIVE_ID}/root:/{FOLDER_NAME}'
    resp = requests.get(url, headers=_headers(token))
    if resp.status_code == 200:
        return resp.json()['id']
    url = f'{GRAPH}/drives/{DRIVE_ID}/root/children'
    resp = requests.post(url, headers=_headers(token), json={'name': FOLDER_NAME, 'folder': {}, '@microsoft.graph.conflictBehavior': 'rename'})
    resp.raise_for_status()
    return resp.json()['id']

def _create_blank_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Registrations'
    ws.append(HEADERS_ROW)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def _get_or_create_excel(token, folder_id):
    url = f'{GRAPH}/drives/{DRIVE_ID}/items/{folder_id}:/{EXCEL_FILENAME}'
    resp = requests.get(url, headers=_headers(token))
    if resp.status_code == 200:
        item_id = resp.json()['id']
        download_url = resp.json()['@microsoft.graph.downloadUrl']
        content = requests.get(download_url).content
        return item_id, content
    blank = _create_blank_workbook()
    upload_url = f'{GRAPH}/drives/{DRIVE_ID}/items/{folder_id}:/{EXCEL_FILENAME}:/content'
    resp = requests.put(upload_url, headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'}, data=blank)
    resp.raise_for_status()
    return resp.json()['id'], blank

def _upload_workbook(token, folder_id, workbook_bytes):
    url = f'{GRAPH}/drives/{DRIVE_ID}/items/{folder_id}:/{EXCEL_FILENAME}:/content'
    requests.put(url, headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'}, data=workbook_bytes).raise_for_status()

def create_client_folder(client_name, pan_number=''):
    if not all([TENANT_ID, CLIENT_ID, CLIENT_SECRET, DRIVE_ID]):
        return None
    safe_name = client_name.strip().replace(' ', '_')
    pan_suffix = pan_number.strip()[-4:].upper() if pan_number else ''
    folder_name = f'{safe_name}_{pan_suffix}' if pan_suffix else safe_name
    try:
        token = _get_token()
        folder_id = _get_or_create_folder(token)
        url = f'{GRAPH}/drives/{DRIVE_ID}/items/{folder_id}/children'
        resp = requests.post(url, headers=_headers(token), json={'name': folder_name, 'folder': {}, '@microsoft.graph.conflictBehavior': 'rename'})
        resp.raise_for_status()
        return resp.json().get('webUrl', '')
    except Exception as e:
        print(f'[OneDrive] Error: {e}')
        return None

def append_registration(row_id, name, dob, pan_number, email, mobile, address, city, notes, submitted_at):
    if not all([TENANT_ID, CLIENT_ID, CLIENT_SECRET, DRIVE_ID]):
        return False
    try:
        token = _get_token()
        folder_id = _get_or_create_folder(token)
        _, wb_bytes = _get_or_create_excel(token, folder_id)
        wb = openpyxl.load_workbook(io.BytesIO(wb_bytes))
        ws = wb.active
        ws.append([row_id, name, str(dob), pan_number, email, mobile, address or '', city or '', notes or '', str(submitted_at)])
        buf = io.BytesIO()
        wb.save(buf)
        _upload_workbook(token, folder_id, buf.getvalue())
        return True
    except Exception as e:
        print(f'[OneDrive] Error: {e}')
        return False
