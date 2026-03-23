"""
OneDrive service for the client portal.

What it does:
  1. Gets an access token from Microsoft using your app credentials.
  2. Finds (or creates) a folder in OneDrive called ONEDRIVE_FOLDER_NAME.
  3. Finds (or creates) an Excel file called 'registrations.xlsx' inside that folder.
  4. Appends a new row for each registration.

Required .env variables:
  ONEDRIVE_TENANT_ID
  ONEDRIVE_CLIENT_ID
  ONEDRIVE_CLIENT_SECRET
  ONEDRIVE_DRIVE_ID
  ONEDRIVE_FOLDER_NAME   (e.g. "Wealixir Registrations")
"""

import io
import requests
import openpyxl
from config.authentication import (
    ONEDRIVE_TENANT_ID     as TENANT_ID,
    ONEDRIVE_CLIENT_ID     as CLIENT_ID,
    ONEDRIVE_CLIENT_SECRET as CLIENT_SECRET,
    ONEDRIVE_DRIVE_ID      as DRIVE_ID,
    ONEDRIVE_FOLDER_NAME   as FOLDER_NAME,
)

EXCEL_FILENAME = 'registrations.xlsx'
HEADERS_ROW    = ['ID', 'Name', 'DOB', 'PAN Number', 'Email', 'Mobile', 'Address', 'City', 'Notes', 'Submitted At']

GRAPH = 'https://graph.microsoft.com/v1.0'


# ---------- Auth ----------

def _get_token():
    url = f'https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token'
    resp = requests.post(url, data={
        'grant_type':    'client_credentials',
        'client_id':     CLIENT_ID,
        'client_secret': CLIENT_SECRET,
        'scope':         'https://graph.microsoft.com/.default',
    })
    resp.raise_for_status()
    return resp.json()['access_token']


def _headers(token):
    return {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}


# ---------- Folder ----------

def _get_or_create_folder(token):
    """Return the OneDrive item ID of the target folder, creating it if needed."""
    # Try to get it
    url = f'{GRAPH}/drives/{DRIVE_ID}/root:/{FOLDER_NAME}'
    resp = requests.get(url, headers=_headers(token))
    if resp.status_code == 200:
        return resp.json()['id']

    # Folder not found — create it
    url = f'{GRAPH}/drives/{DRIVE_ID}/root/children'
    resp = requests.post(url, headers=_headers(token), json={
        'name': FOLDER_NAME,
        'folder': {},
        '@microsoft.graph.conflictBehavior': 'rename',
    })
    resp.raise_for_status()
    return resp.json()['id']


# ---------- Excel file ----------

def _create_blank_workbook():
    """Return bytes of a new .xlsx with just the header row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Registrations'
    ws.append(HEADERS_ROW)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _get_or_create_excel(token, folder_id):
    """Return (item_id, workbook_bytes). Creates the file if it doesn't exist."""
    url = f'{GRAPH}/drives/{DRIVE_ID}/items/{folder_id}:/{EXCEL_FILENAME}'
    resp = requests.get(url, headers=_headers(token))

    if resp.status_code == 200:
        item_id      = resp.json()['id']
        download_url = resp.json()['@microsoft.graph.downloadUrl']
        content      = requests.get(download_url).content
        return item_id, content

    # File doesn't exist — upload a blank one
    blank = _create_blank_workbook()
    upload_url = f'{GRAPH}/drives/{DRIVE_ID}/items/{folder_id}:/{EXCEL_FILENAME}:/content'
    resp = requests.put(
        upload_url,
        headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'},
        data=blank,
    )
    resp.raise_for_status()
    return resp.json()['id'], blank


def _upload_workbook(token, folder_id, workbook_bytes):
    """Upload (overwrite) the Excel file in the folder."""
    url = f'{GRAPH}/drives/{DRIVE_ID}/items/{folder_id}:/{EXCEL_FILENAME}:/content'
    requests.put(
        url,
        headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'},
        data=workbook_bytes,
    ).raise_for_status()


# ---------- Public functions ----------

def create_client_folder(client_name, pan_number=''):
    """
    Create a subfolder inside the main registrations folder.
    Folder name format: Raj_Prasad_234F  (name underscored + last 4 chars of PAN)
    Returns the folder URL on success, None on failure.
    """
    if not all([TENANT_ID, CLIENT_ID, CLIENT_SECRET, DRIVE_ID]):
        return None

    # Build folder name: spaces → underscores, append last 4 of PAN
    safe_name   = client_name.strip().replace(' ', '_')
    pan_suffix  = pan_number.strip()[-4:].upper() if pan_number else ''
    folder_name = f'{safe_name}_{pan_suffix}' if pan_suffix else safe_name

    try:
        token     = _get_token()
        folder_id = _get_or_create_folder(token)

        url  = f'{GRAPH}/drives/{DRIVE_ID}/items/{folder_id}/children'
        resp = requests.post(url, headers=_headers(token), json={
            'name': folder_name,
            'folder': {},
            '@microsoft.graph.conflictBehavior': 'rename',
        })
        resp.raise_for_status()
        web_url = resp.json().get('webUrl', '')
        print(f'[OneDrive] Folder created: {folder_name}')
        return web_url

    except Exception as e:
        print(f'[OneDrive] Folder creation error: {e}')
        return None


def append_registration(row_id, name, dob, pan_number, email, mobile, address, city, notes, submitted_at):
    """
    Append one row to registrations.xlsx in OneDrive.
    Returns True on success, False on failure.
    """
    if not all([TENANT_ID, CLIENT_ID, CLIENT_SECRET, DRIVE_ID]):
        return False   # credentials not configured — skip silently

    try:
        token     = _get_token()
        folder_id = _get_or_create_folder(token)
        _, wb_bytes = _get_or_create_excel(token, folder_id)

        # Add the new row
        wb = openpyxl.load_workbook(io.BytesIO(wb_bytes))
        ws = wb.active
        ws.append([row_id, name, str(dob), pan_number, email, mobile,
                   address or '', city or '', notes or '', str(submitted_at)])

        buf = io.BytesIO()
        wb.save(buf)

        _upload_workbook(token, folder_id, buf.getvalue())
        return True

    except Exception as e:
        # Print detailed error so it's easier to diagnose
        import requests as req_lib
        if isinstance(e, req_lib.exceptions.HTTPError):
            print(f'[OneDrive] HTTP {e.response.status_code} error')
            print(f'[OneDrive] URL: {e.response.url}')
            try:
                print(f'[OneDrive] Response: {e.response.json()}')
            except Exception:
                print(f'[OneDrive] Response: {e.response.text}')
        else:
            print(f'[OneDrive] Error: {e}')
        return False
